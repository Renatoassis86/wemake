"""
Le o arquivo Excel da Regua de Parceria e extrai:
- Nomes das abas
- Estrutura de cada aba (celulas com valores e formulas)
"""
import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl

ARQUIVO = r"C:\Users\Renato\Desktop\projetos_wemake\app_comercial_We Make\Régua_de_Parceria_2026_-_We_Make_v4 (2).xlsm"

wb = openpyxl.load_workbook(ARQUIVO, read_only=False, data_only=False, keep_vba=True)

print("=" * 70)
print("ABAS DO ARQUIVO:")
print("=" * 70)
for name in wb.sheetnames:
    print(f"  - {name}")

print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print()
    print("=" * 70)
    print(f"ABA: {sheet_name}")
    print("=" * 70)

    # Coleta celulas nao-vazias
    cells_data = []
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is None:
                continue
            coord = cell.coordinate
            cells_data.append((coord, val))

    # Mostra as primeiras 120 celulas nao-vazias
    for coord, val in cells_data[:120]:
        val_str = str(val)[:120]
        print(f"  {coord:8s}  {val_str}")

    if len(cells_data) > 120:
        print(f"  ... (+{len(cells_data)-120} celulas omitidas)")

wb.close()
